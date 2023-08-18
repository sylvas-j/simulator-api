from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, View
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from openpyxl import load_workbook
import os
import itertools
from django.utils import timezone

# rest framework
from rest_framework.views import APIView
from rest_framework.response import Response
# from rest_framework.parsers import FormParser, FileUploadParser
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated


# from sklearn.preprocessing import StandardScaler
from court_cases_classification.models import UploadCourtCase
from court_cases_classification.forms import UploadCourtCaseForm
from .court_cases.predict import Predict
import pandas as pd
import csv

from helpers.decorators import unauthenticated_user
# from django.contrib.auth.decorators import login_required


def logoutUser(request):
    logout(request)
    return redirect('home')


class DashboardView(LoginRequiredMixin,TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        # context['cls'] = Lecturer.objects.count()
        # context['results'] = Results.objects.count()
        # context['students'] = Student.objects.count()
        # context['subjects'] = Subject.objects.count()
        return context
    

# def index(request):
#     # template_name = "index.html"
#     if request.method == "POST":
#         username = request.POST['username']
#         password = request.POST['password']
#         print("\nUser Name = ",username)
#         print("Password = ",password)
#         user = authenticate(username=username, password=password)
#         if user is not None:
#             login(request, user)
#             return redirect('smart_home_monitoring:dashboard')
            
#         else:
#             context = {'message':'Invalid User Name and Password'}
#             return render(request, 'index.html', context)
#     return render(request, 'index.html', {'name': 'admin', 'pass': 'Info@123'})


class LoginView(APIView):
    def post(self, request):
        print('post')
        data = self.request.data
        print(data)
        username = data['username']
        password = data['password']
        # images = self.request.data.getlist('images')
        # components = self.request.data.getlist('coms')

        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return Response({'message': 'Login Successful'},
            status=status.HTTP_201_CREATED)
        else:
            context = {'message':'Invalid User Name and Password'}
            return Response(context,status=status.HTTP_201_CREATED)



directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    
file_db = os.path.join(directory, 'court_cases_classification/court_cases/sherloc_court_cases_7.csv')


# from io import BytesIO
class UploaDataView(APIView):
    def post(self, request):
        context = {}
        print('post')
        data = self.request.data
        print(data)
        if len(data) >= 0:
            text = self.request.data['text']
            file = self.request.data['file']
            # print(file)
            # if file == 'undefined':
            #     print('jjjjjjjj')
            # file = request.FILES["file"]
            if file != 'undefined':
                file = request.FILES["file"]
                # for file in files:
                wb = load_workbook(file)
                sheet = wb[wb.sheetnames[0]]
                df = pd.DataFrame(sheet.values,columns=['text'])
                data = itertools.chain.from_iterable(df.values)
            else:
                df = pd.DataFrame({'text':[text]})
                data = itertools.chain.from_iterable(df.values)
            
            X = Predict.transform_text(df)
            predR = Predict.pred_r(X)
            predC = Predict.pred_c(X)

            predRList = []
            for i in predR:
                if i == 0:
                    predRList.append('Not Sentenced')
                else:
                    predRList.append('Sentenced')

            # UploadCourtCaseForm.save(.)
            text = [i for i in data]
            df1 = pd.DataFrame({'text':text,'crime_types':predC,'sentence':predRList})
            # df1 = pd.DataFrame({'crime_types':predC,'sentence':predRList})
            df2 = pd.read_csv(file_db)
            df3 = pd.concat([df2,df1],axis=0)
            df3.to_csv(file_db,index=False)

            object_list = zip(text,predRList,predC)
            # object_list = result
            context = {'message':'- Successfully','data':object_list}
            return Response(context,status=status.HTTP_201_CREATED)

        else:
            context = {'message':'- Form Empty!'}
            return Response(context,status=status.HTTP_201_CREATED)


class DownloadView(APIView):
    def get(self, request):
# def downloads(request):
    # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(
            content_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename=court cases.csv'},
        )
        df = pd.read_csv(file_db)
        writer = csv.writer(response)
        writer.writerow(["text", "crime_types", "sentence"])
        for i in range(len(df)):
            writer.writerow(df.values[i])
        return response





